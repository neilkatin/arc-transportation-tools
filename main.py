#!/usr/bin/env python

import os
import re
import sys
import logging
import argparse
import datetime

import init_logging

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.styles.colors
import dotenv

import config as config_static


log = logging.getLogger(__name__)


DATESTAMP = datetime.datetime.now().strftime("%Y-%m-%d")

class AttrDict(dict):
    def __init__(self, *args, **kwargs):
        super(AttrDict, self).__init__(*args, **kwargs)
        self.__dict__ = self

def rentals_filter(row):
    return row['Ctg'] == 'R' and row['Key'] == ''

def empty_filter(row):
    return True

def current_filter(row):
    return row['Ctg'] == 'R' and row['Status'] == 'Active'


def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    config_dotenv = dotenv.dotenv_values(verbose=True)

    config = AttrDict()
    for item in dir(config_static):
        if not item.startswith('__'):
            config[item] = getattr(config_static, item)

    #log.debug(f"config after copy: { config.keys() }")

    for key, val in config_dotenv.items():
        config[key] = val

    vehicles_file = config.VEHICLES
    if not os.path.exists(vehicles_file):
        log.fatal(f"Vehicles file { vehicles_file } not found")
        sys.exit(1)

    vehicles_wb = openpyxl.load_workbook(vehicles_file)
    vehicles_ws = vehicles_wb[config.VEHICLES_SHEET_NAME]
    vehicles_map = build_map(vehicles_ws, "Vehicles", 1, ["Rcvd From"], rentals_filter)

    # delete workbook if it exists
    output_file = config.OUTPUT_WB
    if os.path.exists(output_file):
        # eventually copy fields out...
        os.remove(output_file)

    # create a new one
    output_wb = openpyxl.Workbook()

    staff_file = config.STAFF_ROSTER
    staff_map = None
    current_ws = None
    vehicles_spec = [ ['Name', 20 ], ['Reservation No', 15 ], [ 'GAP', 15 ], ['Date Received', 12 ] ]
    vehicles_spec_len = len(vehicles_spec)
    if not os.path.exists(staff_file):
        log.info(f"skipping { config.MERGED_SHEET_NAME } sheet: could not find staff roster { staff_file }")
    else:

        log.debug(f"generating { config.MERGED_SHEET_NAME } sheet using { staff_file }")
        staff_wb = openpyxl.load_workbook(staff_file)
        staff_ws = staff_wb[config.STAFF_ROSTER_SHEET_NAME]
        staff_map = build_map(staff_ws, "Staff Roster", config.STAFF_ROSTER_TITLE_ROW, ["Name"], empty_filter)

        # handle the merged sheet
        merged_ws = output_wb.create_sheet(title=config.MERGED_SHEET_NAME)
        merged_ws.freeze_panes = merged_ws['B2']

        staff_spec = [ [ 'Email', 30 ], [ 'Cell phone', 14 ], [ 'Assigned', 20 ], [ 'Checked in', 20 ], [ 'Supervisor(s)', 20 ] ]

        vehicles_spec.append([ 'Make', 10 ])
        vehicles_spec.append([ 'Model', 10 ])
        vehicles_spec.append([ 'Color', 10 ])
        vehicles_spec.append([ 'Key', 10 ])
        vehicles_spec.append([ 'Plate', 10 ])
        vehicles_spec.append([ 'Tag', 4 ])
        vehicles_spec.append([ 'Driver', 20 ])

        make_merged(merged_ws, vehicles_map, vehicles_spec, staff_map, staff_spec)

        # handle the active sheet
        current_ws = output_wb.create_sheet(title=config.CURRENT_SHEET_NAME)
        current_ws.freeze_panes = current_ws['B2']
        vehicles_map = build_map(vehicles_ws, "Vehicles", 1, ["Driver","Rcvd From"], current_filter)
        make_merged(current_ws, vehicles_map, vehicles_spec, staff_map, staff_spec)


    outprocessed_file = config.OUTPROCESSED_ROSTER
    if not os.path.exists(outprocessed_file):
        log.info(f"skipping { config.OUTPROCESSED_SHEET_NAME } sheet: could not find outprocessed roster file { outprocessed_file }")
    else:
        log.debug(f"generating { config.OUTPROCESSED_SHEET_NAME } sheet using { outprocessed_file }")

        outroster_wb = openpyxl.load_workbook(outprocessed_file)
        outroster_ws = outroster_wb[config.OUTPROCESSED_ROSTER_SHEET_NAME]
        outroster_map = build_map(outroster_ws, "Outprocessed Roster", config.OUTPROCESSED_ROSTER_TITLE_ROW, ["Name"], empty_filter)
        outroster_spec = [ [ 'Email', 30 ], [ 'Cell phone', 14 ], [ 'Checked in', 20 ], [ 'Released', 20 ], [ 'Supervisor(s)', 20 ] ]

        out_ws = output_wb.create_sheet(title=config.OUTPROCESSED_SHEET_NAME)
        out_ws.freeze_panes = out_ws['B2']
        vehicles_map = build_map(vehicles_ws, "Vehicles", 1, ["Driver","Rcvd From"], current_filter)
        make_merged(out_ws, vehicles_map, vehicles_spec, outroster_map, outroster_spec, suppress_missing=True)


        if staff_map != None:
            out_ws = output_wb.create_sheet(title=config.MISSING_SHEET_NAME)
            left_map, right_map = filter_tables(vehicles_map, staff_map, filter_left_only)
            make_merged(out_ws, left_map, vehicles_spec, outroster_map, outroster_spec)

    # handle reconciliation; ok if open_rentals_files isn't there; just don't make the sheet if it isn't
    open_file = config.OPEN_RENTALS
    open_rentals_ws = None
    closed_rentals_ws = None
    if not os.path.exists(open_file):
        log.info(f"skipping { config.RECONCILED_SHEET_NAME } sheet: could not find avis file { open_file }")
    else:
        log.debug(f"generating { config.RECONCILED_SHEET_NAME } sheet using { open_file }")
        rentals_wb = openpyxl.load_workbook(open_file, read_only=False)
        open_rentals_ws = rentals_wb[config.OPEN_RENTALS_SHEET_NAME]
        #open_rentals_ws.reset_dimensions()      # useful for read-only spreadsheets

        reconciled_ws = output_wb.create_sheet(title=config.RECONCILED_SHEET_NAME)
        reconciled_ws.freeze_panes = reconciled_ws['B2']

        make_reconciled(reconciled_ws, open_rentals_ws, config.OPEN_RENTALS_TITLE_ROW, vehicles_ws, 1, config.OPEN_RENTALS_DRS)

    # if neither sheet was created: give an error
    if len(output_wb.sheetnames) == 1:
        log.fatal(f"Neither the AVIS file ({ open_file }) nor the staff roster ({ staff_file }) were present.  Aborting...")
        sys.exit(1)

    if current_ws is not None:
        annotate_vehicles_with_avis(current_ws, vehicles_spec_len, open_rentals_ws, config.OPEN_RENTALS_TITLE_ROW,
                closed_rentals_ws, config.CLOSED_RENTALS_TITLE_ROW)


    # delete the initial default sheet
    default_sheet_name = 'Sheet'
    if default_sheet_name in output_wb:
        del output_wb[default_sheet_name]

    # save the file
    output_wb.save(output_file)



def make_merged(out_ws, vehicles_map, vehicles_spec, staff_map, staff_spec, suppress_missing=False):
    """ generate the merged sheet from vehicles and staff

        The vehicles_spec and staff_spec control which Columns from the source worksheet are output.
        Each entry in the array is a 2 element tuple of [ 'Column Name', column_width ].

        suppress_missing means: don't output the line if there is no matching join in the staff_map
    """

    #log.debug(f"make_merged: vehicles_map size: { len(vehicles_map) }")

    vehicles_columns = []
    vehicles_col_width = []
    for e in vehicles_spec:
        vehicles_columns.append(e[0])
        vehicles_col_width.append(e[1])

    staff_columns = []
    staff_col_width = []
    for e in staff_spec:
        staff_columns.append(e[0])
        staff_col_width.append(e[1])

    column_map = {}
    column_index = 0
    col_dims = out_ws.column_dimensions
    for i, name in enumerate(vehicles_columns):
        column_index += 1
        column_map[name] = { 'index': column_index, 'map': vehicles_map, 'col_name': name, 'map_name': 'Vehicles' }
        out_ws.cell(row=1, column=column_index, value=name)
        col_dims[openpyxl.utils.get_column_letter(column_index)].width = vehicles_col_width[i]

    for i, name in enumerate(staff_columns):
        column_index += 1
        column_map[name] = column_index
        column_map[name] = { 'index': column_index, 'map': staff_map, 'col_name': name, 'map_name': 'Roster' }
        out_ws.cell(row=1, column=column_index, value=name)
        col_dims[openpyxl.utils.get_column_letter(column_index)].width = staff_col_width[i]

    # the name in the vehicles_map is 'Rcvd From'
    column_map['Name']['col_name'] = 'Rcvd From'

    # now generate the data
    keys = sorted(vehicles_map.keys(), key=str.lower)

    rownum = 1
    in_rownum = 1
    for row_name in keys:
        row = vehicles_map[row_name]
        failed_row = False
        in_rownum += 1

        if suppress_missing:
            # skip the row if not matching columns in roster map
            found_row = False
            for name, data in column_map.items():
                out_index = data['index']
                in_map = data['map']
                in_col = data['col_name']

                if in_map != staff_map:
                    continue

                if row_name in in_map:
                    found_row = True

            if not found_row:
                #log.debug(f"Ignoring input row { in_rownum }: no matches on roster")
                continue

        rownum += 1
        for name, data in column_map.items():
            out_index = data['index']
            in_map = data['map']
            in_col = data['col_name']

            if row_name not in in_map:
                continue

            entry = in_map[row_name]

            if in_col not in entry:
                log.error(f"can't find key '{ in_col }' in map '{ data['map_name'] }'")
            else:
                value = entry[in_col]
                out_ws.cell(row=rownum, column=out_index, value=value)


def filter_tables(left_map, right_map, filter_func):
    """ filter rows in specified dictionaries with a given filter function

        (part of refactoring: separating row selection from output generation.)

        This function joins the left and right map, then passes the individual entries
        to the filter function.  If the function returns False then the entry is deleted
        from the output maps.

        The input maps are not changed; new left and right maps are returned.

        The two maps are assumed to be keyed the same.
    """
    left_out = {}
    right_out = {}

    for key in left_map.keys():
        left_entry = left_map[key]
        if key in right_map:
            right_entry = right_map[key]
        else:
            right_entry = None

        if filter_func(left_entry, right_entry):
            left_out[key] = left_entry
            if right_entry is not None:
                right_out[key] = right_entry

    return left_out, right_out

def filter_both(left, right):
    """ return true if both left and right exist """
    return left is not None and right is not None

def filter_left_only(left, right):
    """ return true if both left and right exist """
    return left is not None and right is None




def make_reconciled(reconciled_ws, rentals_ws, rentals_starting_row, vehicles_ws, vehicles_starting_row, dr_list):
    """ generate reconciled ws from rentals_ws, marking which key numbers and reservation numbers are in vehicles_ws """
    
    rentals_name_map, rentals_cols = process_title_row(rentals_ws, rentals_starting_row)
    vehicles_name_map, vehicles_cols = process_title_row(vehicles_ws, vehicles_starting_row)

    #log.debug(f"rentals_name_map: { rentals_name_map }")

    # go through the vehicles sheet and gather up all the reservation and key numbers.
    key_map = gather_column(vehicles_ws, vehicles_name_map['Key'], vehicles_starting_row, "Vehicles", 'Key')
    reservation_map = gather_column(vehicles_ws, vehicles_name_map['Reservation No'], vehicles_starting_row, "Vehicles", 'Reservation No')
    plate_map = gather_column(vehicles_ws, vehicles_name_map['Plate'], vehicles_starting_row, "Vehicles", 'Plate')

    # cell colors we'll be using
    fill_red = openpyxl.styles.PatternFill(fgColor="FFC0C0", fill_type = "solid")
    fill_green = openpyxl.styles.PatternFill(fgColor="C0FFC0", fill_type = "solid")
    fill_yellow = openpyxl.styles.PatternFill(fgColor="FFFFC0", fill_type = "solid")
    fill_blue = openpyxl.styles.PatternFill(fgColor="C0C0FF", fill_type = "solid")

    plate_column = rentals_name_map['License Plate Number']
    key_column = rentals_name_map['MVA No']
    res_column = rentals_name_map['Reservation No']
    drs_column = rentals_name_map['Cost Control No']
    match_map = { key_column: key_map, res_column: reservation_map, plate_column: plate_map }
    match_fixups = { key_column: lambda x: re.sub('^0','',x), res_column: lambda x: re.sub('[-]','',x), plate_column: lambda x: x }

    log.debug("before output generation")

    # turn the list of drs into a dict for easy matching
    dr_map = {}
    for item in dr_list:
        dr_map[item] = 1

    column_dims = reconciled_ws.column_dimensions

    column_widths = {
            'Rental Region Desc': 15,
            'Rental Zone Desc': 15,
            'Rental Distict Desc': 10,
            'MVA No': 10,
            'License Plate State Code': 5,
            'License Plate Number': 10,
            'Make': 6,
            'Model': 6,
            'Ext Color Code': 6,
            'Reservation No': 15,
            'Rental Agreement No': 15,
            'CO Date': 12,
            'CO Time': 10,
            'Rental Loc Mnemonic': 15,
            'Address Line 1': 10,
            'Address Line 3': 10,
            'Full Name': 20,
            'Return Loc Mnomonic': 5,
            'Exp CI Loc Id': 5,
            'Exp CI Date': 12,
            'Exp CI Time': 10,
            'AWD Orgn Buildup Desc': 5,
            'Cost Control No': 10,
            'Booking Source Emp no': 10
            }

    for column_name in column_widths:
        if column_name in rentals_name_map:
            column_width = column_widths[column_name]
            column_index = rentals_name_map[column_name]
            # -1 in column index is because we're shifting all columns to the left since input has no column A
            column_dims[openpyxl.utils.get_column_letter(column_index-1)].width = column_width

    # specify date columns
    date_columns = [ 'CO Date', 'Exp CI Date' ]
    date_column_map = {}
    for name in date_columns:
        date_column_map[name] = 1


    # now mark the reconciled_ws with the right colors for items found/not found in vehicles
    output_row = 0
    for row in rentals_ws.iter_rows(min_row=rentals_starting_row, values_only=False):

        dr = row[drs_column -1].value

        # filter out unmatched DRs from list
        if output_row != 0 and dr not in dr_map:
            continue

        #log.debug(f"output generation row { output_row } dr { dr }")

        output_row += 1
        for cell in row:
            # skip blank first column
            if cell.column == 1:
                continue

            cell_title = rentals_cols[cell.column]

            #log.debug(f"setting row { output_row } column { cell.column } to '{ cell.value }'")

            # adjust output column by 1 because input sheet has no column 'A'
            input_column = cell.column
            output_column = input_column -1
            value = cell.value
            out_cell = reconciled_ws.cell(row=output_row, column=output_column, value=value)

            # ignore title row
            if output_row == 1:
                continue

            # make date columns look like dates
            if cell_title in date_column_map:
                out_cell.number_format = 'yyyy-mm-dd'

        # now fix colors, but not for title row
        if output_row != 1:

            match_array = []
            for col, cmap in match_map.items():
                rental_value = row[col -1].value

                # need to clean up data to cannonicalize it
                if isinstance(rental_value, int):
                    rental_value = str(rental_value)
                rental_value = match_fixups[col](rental_value)
                if rental_value in cmap:
                    match_row = cmap[rental_value]
                else:
                    match_row = None
                match_array.append(match_row)
                #log.debug(f"match: row { output_row } col '{ col }' rental_value '{ rental_value }' match_row '{ match_row }'")

            # if all the entries match: mark them green
            if match_array.count(None) == len(match_array):
                for col in match_map.keys():
                    reconciled_ws.cell(row=output_row, column=col-1).fill = fill_blue

            elif match_array.count(match_array[0]) == len(match_array):
                # all match - mark them all green
                for col in match_map.keys():
                    reconciled_ws.cell(row=output_row, column=col-1).fill = fill_green
            else:
                index = 0
                for col, cmap in match_map.items():
                    value = match_array[index]
                    index += 1
                    if value == None:
                        reconciled_ws.cell(row=output_row, column=col-1).fill = fill_red
                    else:
                        reconciled_ws.cell(row=output_row, column=col-1).fill = fill_yellow

        # DEBUG ONLY
        #if output_row > 10:
        #    break




def gather_column(ws, column_num, title_row_num, table_name, column_name):
    """ gather all the values in a column into dict for easy matching """

    results = {}
    row_num = title_row_num
    for row in ws.iter_rows(min_row=title_row_num+1, min_col=column_num, max_col=column_num, values_only=True):
        row_num += 1
        cell = row[0]
        if cell == '':
            continue

        # delete formatting characters
        if isinstance(cell, str):
            cell = re.sub('[ \t-]', '', cell)

        value = str(cell)
        if value in results:

            # horrible hack to deal with n/a column in reservation no: just ignore values of n/a
            if value != 'N/A':
                # this is a duplicate entry
                log.error(f"Found duplicate for table { table_name } column { column_name }: old row { results[value] }, new row { row_num }")

        results[value] = row_num

    return results

def process_title_row(sheet, title_row_num):
    """ build two maps: one mapping column name to column number, and one from column number to name """
    title_name_map = {}
    title_cols = {}

    for row in sheet.iter_rows(min_row=title_row_num, max_row=title_row_num, values_only=False):
        for cell in row:
            #log.debug(f"title { cell.column }, { cell.row } = '{ cell.value }'")
            title_name_map[cell.value] = cell.column
            title_cols[cell.column] = cell.value
        break

    return title_name_map, title_cols

def build_map(sheet, sheet_name, starting_row, key_name_list, row_filter):
    # grab the title row
    title_name_map, title_cols = process_title_row(sheet, starting_row)

    #log.debug(f"title_name_map: { title_name_map }")

    results = {}
    row_num = starting_row
    for row in sheet.iter_rows(min_row=starting_row+1, values_only=False):
        row_num += 1
        row_map = { 'row_num': row_num, 'sheet_name': sheet_name }
        for cell in row:
            value = cell.value
            title = title_cols[cell.column]
            row_map[title] = value

        if not row_filter(row_map):
            continue

        #log.debug(f"row_map: { row_map }")
        for name in key_name_list:
            key = row_map[name]
            if key != '' and key != None:
                break

        if key == '' or key == None:
            log.error(f"build_map: empty key from { key_name_list } for row { row_num }")

        if key in results:
            log.error(f"Error: duplicate entry for entry { key } on row { row_num } and { results[key]['row_num'] }")

        results[key] = row_map

    return results


def annotate_vehicles_with_avis(current_ws, insert_col, open_ws, open_starting_row, closed_ws, closed_starting_row):

    current_ws.insert_cols(insert_col + 1)
    current_ws.cell(row=1, column=insert_col + 1, value="Avis")

    #for i in range(1,  current_ws.max_column +1):
    #    log.debug(f"after insert column { openpyxl.utils.get_column_letter(i) } to { col_dims[openpyxl.utils.get_column_letter(i)].width }")

    # inserting a column doesn't shift column widths; do so
    col_dims = current_ws.column_dimensions
    for i in range(current_ws.max_column + 1, insert_col + 1, -1):
        col_dims[openpyxl.utils.get_column_letter(i)].width = col_dims[openpyxl.utils.get_column_letter(i -1)].width
        #log.debug(f"set column { openpyxl.utils.get_column_letter(i) } to { col_dims[openpyxl.utils.get_column_letter(i)].width }")

    current_name_map, current_cols = process_title_row(current_ws, 1)

    if open_ws is not None:
        open_map, open_cols = process_title_row(open_ws, open_starting_row)

        






def parse_args():
    parser = argparse.ArgumentParser(
            description="process support for the regional bootcamp mission card system",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")

    #group = parser.add_mutually_exclusive_group(required=True)
    #group.add_argument("-p", "--prod", "--production", help="use production settings", action="store_true")
    #group.add_argument("-d", "--dev", "--development", help="use development settings", action="store_true")

    args = parser.parse_args()
    return args


if __name__ == "__main__":
    main()

